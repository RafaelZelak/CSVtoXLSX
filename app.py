import os
import pandas as pd
from tkinter import messagebox
import customtkinter as ctk
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import re

# Configurações globais do tema
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

def extract_socio_info(socio_data):
    """Extrai informações dos sócios e retorna como uma lista de dicionários."""
    if not isinstance(socio_data, str):
        return []
    socios = []
    socio_patterns = re.findall(r"Nome: (.*?), Faixa Etária: (.*?), Qualificação: (.*?), Data Entrada: (.*?)(;|$)", socio_data)
    for i, match in enumerate(socio_patterns, start=1):
        socio_info = {
            f"Nome Sócio {i}": match[0],
            f"Faixa Etária Sócio {i}": match[1],
            f"Qualificação Sócio {i}": match[2],
            f"Data Entrada Sócio {i}": match[3]
        }
        socios.append(socio_info)
    return socios

def extract_business_hours(business_hours):
    """Extrai os horários de funcionamento e retorna um dicionário com os dias da semana."""
    if not isinstance(business_hours, str):
        return {day: 'Fechado' for day in ['segunda-feira', 'terça-feira', 'quarta-feira', 'quinta-feira', 'sexta-feira', 'sábado', 'domingo']}
    days = ['segunda-feira', 'terça-feira', 'quarta-feira', 'quinta-feira', 'sexta-feira', 'sábado', 'domingo']
    business_hours_dict = {day: 'Fechado' for day in days}
    business_hours_patterns = re.findall(r"(\w+-feira): (.*?)(;|$)", business_hours)
    for match in business_hours_patterns:
        day, hours = match[0], match[1]
        if day in business_hours_dict:
            business_hours_dict[day] = hours.strip()
    return business_hours_dict

def process_csv_to_excel(csv_file_path, output_file_path):
    df = pd.read_csv(csv_file_path, dtype={'CNPJ': str})

    # Extração de informações de sócios
    if 'Sócios (Nome, Faixa Etária, Qualificação, Data Entrada)' in df.columns:
        socio_column = 'Sócios (Nome, Faixa Etária, Qualificação, Data Entrada)'
        new_columns = []
        for idx, row in df.iterrows():
            socio_data = row[socio_column]
            socios_info = extract_socio_info(socio_data)
            for socio in socios_info:
                for key, value in socio.items():
                    if key not in new_columns:
                        df[key] = None
                        new_columns.append(key)
                    df.at[idx, key] = value
        df = df.drop(columns=[socio_column])

    # Extração de horários de funcionamento
    if 'Horários de Funcionamento' in df.columns:
        business_hours_column = 'Horários de Funcionamento'
        new_columns = []
        for idx, row in df.iterrows():
            business_hours_data = row[business_hours_column]
            business_hours_info = extract_business_hours(business_hours_data)
            for day, hours in business_hours_info.items():
                if day not in new_columns:
                    df[day] = None
                    new_columns.append(day)
                df.at[idx, day] = hours
        df = df.drop(columns=[business_hours_column])

    # Ordenar colunas dos sócios para o final
    socio_columns = [col for col in df.columns if "Sócio" in col]
    non_socio_columns = [col for col in df.columns if col not in socio_columns]
    df = df[non_socio_columns + socio_columns]

    # Criação do Excel com cores e estilos
    wb = Workbook()
    ws = wb.active

    contact_group = ['Telefone 1', 'Telefone 2', 'Telefone Enriquecido', 'Email', 'Email Enriquecido']
    address_group = ['Logradouro', 'Município', 'UF', 'CEP', 'Logradouro Enriquecido']
    contact_fill = PatternFill(start_color='FFCCE5', end_color='FFCCE5', fill_type='solid')
    address_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
    socio_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    light_gray_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    ws.append(df.columns.tolist())
    for i, cell in enumerate(ws[1], 1):
        if cell.value in contact_group:
            cell.fill = contact_fill
        elif cell.value in address_group:
            cell.fill = address_fill
        elif 'Sócio' in cell.value:
            cell.fill = socio_fill
        else:
            cell.fill = white_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for index, row in df.iterrows():
        ws.append(row.tolist())
        fill = light_gray_fill if index % 2 == 0 else white_fill
        for cell in ws[index + 2]:
            cell.fill = fill

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = max_length

    wb.save(output_file_path)

def validate_csv_format(csv_file_path):
    """Valida se o arquivo CSV contém as colunas necessárias."""
    required_columns = ['CNPJ', 'Sócios (Nome, Faixa Etária, Qualificação, Data Entrada)', 'Horários de Funcionamento']  # Adicione aqui as colunas obrigatórias
    try:
        df = pd.read_csv(csv_file_path, nrows=1)  # Lê apenas a primeira linha para verificar as colunas
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao ler o arquivo CSV: {e}")
        return False

    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        messagebox.showerror("Erro", f"Colunas ausentes no CSV: {', '.join(missing_columns)}")
        return False

    return True

def select_csv_file():
    csv_file_path = ctk.filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    if csv_file_path:
        csv_entry.delete(0, 'end')
        csv_entry.insert(0, csv_file_path)

def select_output_folder():
    output_directory = ctk.filedialog.askdirectory()
    if output_directory:
        folder_entry.delete(0, 'end')
        folder_entry.insert(0, output_directory)

def convert_file():
    csv_file_path = csv_entry.get()
    output_directory = folder_entry.get()
    if csv_file_path and output_directory:
        # Validação do arquivo CSV antes da conversão
        if not validate_csv_format(csv_file_path):
            return  # Se o CSV não for válido, retorna e espera um novo arquivo

        output_file_name = entry.get() + ".xlsx"
        output_file_path = os.path.join(output_directory, output_file_name)
        process_csv_to_excel(csv_file_path, output_file_path)
        messagebox.showinfo("Sucesso", f"Arquivo Excel criado: {output_file_path}")
    else:
        messagebox.showerror("Erro", "Selecione um arquivo CSV e uma pasta de saída.")

# Criação da janela principal
root = ctk.CTk()
root.title("CSV to Excel Converter")
root.geometry("400x350")

# Ajustes de espaçamento, fonte e padding
padding_y = 10
padding_x = 20
button_width = 220
height = 40
entry_width = 250

# Criação do título
title_label = ctk.CTkLabel(root, text="CSV para Excel", font=("Arial", 18))
title_label.pack(pady=padding_y)

# Frame para seleção de arquivo CSV
csv_frame = ctk.CTkFrame(root)
csv_frame.pack(pady=padding_y, padx=padding_x, fill="x")

# Entry para caminho do arquivo CSV
csv_entry = ctk.CTkEntry(csv_frame, font=("Arial", 12), width=entry_width, height=height)
csv_entry.pack(side="left", padx=(0, 10), expand=True, fill="x")

# Botão para selecionar arquivo CSV
csv_button = ctk.CTkButton(csv_frame, text="Procurar CSV", command=select_csv_file, width=button_width, height=height)
csv_button.pack(side="left")

# Frame para seleção da pasta de saída
folder_frame = ctk.CTkFrame(root)
folder_frame.pack(pady=padding_y, padx=padding_x, fill="x")

# Entry para caminho da pasta de saída
folder_entry = ctk.CTkEntry(folder_frame, font=("Arial", 12), width=entry_width, height=height)
folder_entry.pack(side="left", padx=(0, 10), expand=True, fill="x")

# Botão para selecionar pasta de saída
folder_button = ctk.CTkButton(folder_frame, text="Escolher Pasta", command=select_output_folder, width=button_width, height=height)
folder_button.pack(side="left")

# Nome do arquivo Excel
entry_label = ctk.CTkLabel(root, text="Nome do Arquivo Excel:", font=("Arial", 16))
entry_label.pack(pady=padding_y)
entry = ctk.CTkEntry(root, font=("Arial", 12), width=entry_width, height=height)
entry.pack(pady=padding_y, padx=padding_x, fill="x")

# Botão para converter o arquivo
convert_button = ctk.CTkButton(root, text="Converter", command=convert_file, width=button_width, height=height)
convert_button.pack(pady=padding_y)

root.mainloop()
